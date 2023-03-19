# Challenge #1
# Create a Python script that uses openpyxl and generates an Excel file with a 
# single sheet and the below content:

import openpyxl
from openpyxl.styles import *
import csv

wb = openpyxl.Workbook()
sheet = wb.active

content = {'Year':'Sales', 2017:150000, 2018:180000, 2019:210000, 2020:125000}
for pair in content.items():
    sheet.append(pair)
wb.save('python/challenges/files/data.xlsx')

# Challenge #2
# Change the solution from the previous challenge and set the sheet title 
# to COMPANY SALES.

sheet.title = 'COMPANY SALES'
wb.save('python/challenges/files/data.xlsx')

# Challenge #3
# a) Add a new sheet called VAT.
# b) Read the Excel file in Python, calculate the VAT if the VAT rate is 15%, 
# and write the data into the new sheet.
# c) Save the Excel file as a new file called sales_and_vat.xlsx

wb = openpyxl.load_workbook('python/challenges/files/data.xlsx')
sheet = wb.active
items = list()
for row in sheet.values:
    items.append(row)

vat = list()
for row in items[1:]:
    element = (row[0], row[1] * 0.15)
    vat.append(element)

wb.create_sheet('VAT')
sheet = wb['VAT']
sheet['A1'] = 'Year'
sheet['B1'] = 'VAT'
for row in vat:
    sheet.append(row)
wb.save('python/challenges/files/data2.xlsx')

# Challenge #4
# Add an Excel formula into the B6 cell so that it calculates the total sales.
# Add the string Total Sales into cell C6.

wb = openpyxl.load_workbook('python/challenges/files/data.xlsx')
sheet = wb.active
sheet['B6'].value = '=sum(b2:b5)'
sheet['C6'].value = 'Total Sales'

wb.save('python/challenges/files/data3.xlsx')

# Challenge #5
# Format the cells B6 and C6 in the following way: 
# Tahoma font family, size 16, color red, and bold.

font = Font(name='Tahoma',size=16, color=colors.BLUE, bold=True)
sheet['B6'].font = font
sheet['C6'].font = font
wb.save('python/challenges/files/data3.xlsx')

# Challenge #6
# # Create a Python script that contains a function called csv2excel() which 
# exports any CSV file to an excel one.
# The function will be called like this: csv2excel('people3.csv', 'teachers.xlsx')
# It will read the CSV file called people3.csv and export it to teacher.xlsx

def csv2excel(source, destination):
    lst = []
    with open(source) as f:
        reader = csv.reader(f)
        for row in reader:
            lst.append(row)
    wb = openpyxl.Workbook()
    sheet = wb.active
    for row in lst:
        sheet.append(row)
    wb.save(destination)
csv2excel('python/challenges/files/people3.csv', 'python/challenges/files/teachers.xlsx')

# Challenge #7

def csv2excel(source, destination, delimiter=','):
    lst = []
    with open(source) as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            lst.append(row)
    wb = openpyxl.Workbook()
    sheet = wb.active
    for row in lst:
        sheet.append(row)
    wb.save(destination)
csv2excel('python/challenges/files/people3.csv', 'python/challenges/files/teachers2.xlsx', ':')

# Challenge #8

def excel2csv(source, destination):
    wb = openpyxl.load_workbook(source)
    sheet = wb.active
    with open(destination, 'w') as f:
        writer = csv.writer(f, lineterminator='\n')
        for row in sheet.values:
            writer.writerow(row)

excel2csv('python/challenges/files/books.xlsx', 'python/challenges/files/booklist.csv')
